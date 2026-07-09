"""
ERCOT COP ONRUC Status Pull
============================
Pulls all resource hours with status ONRUC from the ERCOT Public API
using the NP1-301 (60-Day COP Adjustment Period Snapshot) endpoint.
Captures: resource name, operating date, hour ending, HSL, QSE, and status.

SETUP
-----
1. Register at https://apiexplorer.ercot.com/ (free, email-verified)
2. Subscribe to a product on the Products page to get your subscription key
3. Set the three credentials below (or export as env variables)

USAGE
-----
    pip install requests pandas openpyxl
    python ercot_onruc_pull.py

OUTPUT
------
    ercot_onruc_2023_present.csv   — raw CSV of all ONRUC records since 2023-01-01
    ercot_onruc_2023_present.xlsx  — formatted Excel with Summary + one tab per month
"""

import os
import time
import requests
import pandas as pd
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ─── CREDENTIALS ──────────────────────────────────────────────────────────────
USERNAME         = os.getenv("ERCOT_USERNAME",         "YOUR_EMAIL")
PASSWORD         = os.getenv("ERCOT_PASSWORD",         "YOUR_PASSWORD")
SUBSCRIPTION_KEY = os.getenv("ERCOT_SUBSCRIPTION_KEY", "YOUR_SUBSCRIPTION_KEY")

# ─── CONFIG ───────────────────────────────────────────────────────────────────
BASE_URL      = "https://api.ercot.com/api/public-reports"
AUTH_URL      = (
    "https://ercotb2c.b2clogin.com/ercotb2c.onmicrosoft.com"
    "/B2C_1_PUBAPI-ROPC-FLOW/oauth2/v2.0/token"
)
ENDPOINT      = "/np1-301/60_cop_adj_period_snapshot"
CLIENT_ID     = "fec253ea-0d06-4272-a5e6-b478baeecd70"
SCOPE         = f"openid {CLIENT_ID} offline_access"

START_DATE    = "2023-01-01"
END_DATE      = datetime.now(timezone.utc).strftime("%Y-%m-%d")

PAGE_SIZE     = 10000
SLEEP_SECONDS = 0.3
OUTPUT_CSV    = "ercot_onruc_2023_present.csv"
OUTPUT_XLSX   = "ercot_onruc_2023_present.xlsx"

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def get_id_token():
    params = {
        "username":      USERNAME,
        "password":      PASSWORD,
        "grant_type":    "password",
        "scope":         SCOPE,
        "client_id":     CLIENT_ID,
        "response_type": "id_token",
    }
    resp = requests.post(AUTH_URL, data=params, timeout=30)
    resp.raise_for_status()
    token = resp.json().get("id_token")
    if not token:
        raise ValueError(f"Auth failed — no id_token in response: {resp.text[:500]}")
    print("  Authenticated successfully.")
    return token


def make_headers(token):
    return {
        "Authorization":            f"Bearer {token}",
        "Ocp-Apim-Subscription-Key": SUBSCRIPTION_KEY,
        "Accept":                   "application/json",
    }


# ─── FETCH ────────────────────────────────────────────────────────────────────

def fetch_page(headers, page, start, end):
    url = f"{BASE_URL}{ENDPOINT}"
    params = {
        "deliveryDateFrom": start,
        "deliveryDateTo":   end,
        "status":           "ONRUC",
        "size":             PAGE_SIZE,
        "page":             page,
    }
    resp = requests.get(url, headers=headers, params=params, timeout=60)
    if resp.status_code == 401:
        raise PermissionError("401 Unauthorized — check credentials and subscription key.")
    resp.raise_for_status()
    return resp.json()


def fetch_all_onruc(token):
    headers  = make_headers(token)
    all_rows = []
    page     = 1
    total    = None

    print(f"\nFetching COP snapshot: {START_DATE} → {END_DATE}")
    print("Filtering to ONRUC status only ...\n")

    while True:
        print(f"  Page {page}" + (f" / {-(-total // PAGE_SIZE)}" if total else "") + " ...", end=" ")
        data = fetch_page(headers, page, START_DATE, END_DATE)

        meta   = data.get("_meta", {})
        fields = data.get("fields", [])
        rows   = data.get("data", [])

        if fields and isinstance(fields[0], dict):
            fields = [f["name"] for f in fields]

        if total is None:
            total = meta.get("totalRecords", 0)
            print(f"Total ONRUC records in range: {total:,}")

        if not rows:
            print("  No more rows.")
            break

        for row in rows:
            all_rows.append(dict(zip(fields, row)))

        print(f"  → {len(rows):,} on this page | total so far: {len(all_rows):,}")

        if page * PAGE_SIZE >= total or len(rows) < PAGE_SIZE:
            break

        page += 1
        time.sleep(SLEEP_SECONDS)

    if not all_rows:
        print("\nNo ONRUC records found in the date range.")
        return pd.DataFrame()

    return pd.DataFrame(all_rows)


# ─── PROCESS ──────────────────────────────────────────────────────────────────

def build_result(df):
    df.columns = [c.lower() for c in df.columns]

    col_map = {
        "resource":      ["resourcename"],
        "qse":           ["qsename"],
        "operatingdate": ["deliverydate"],
        "hourending":    ["hourending"],
        "hsl":           ["highsustainedlimit"],
        "status":        ["status"],
    }

    out = {}
    for target, candidates in col_map.items():
        for c in candidates:
            if c in df.columns:
                out[target] = df[c]
                break
        if target not in out:
            out[target] = None

    result = pd.DataFrame(out)
    result["hsl"] = pd.to_numeric(result["hsl"], errors="coerce")
    result["operatingdate"] = pd.to_datetime(result["operatingdate"])

    # Parse HE from hourending (hour value IS the HE number: 01:00 = HE1, 24:00 = HE24)
    def time_to_he(t):
        try:
            h = int(str(t).split(":")[0])
            return h if h > 0 else 24
        except:
            return None

    result["he"] = result["hourending"].apply(time_to_he)

    result = result.sort_values(["operatingdate", "he", "resource"]).reset_index(drop=True)
    return result


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

# Styles
NAVY  = PatternFill("solid", fgColor="0D1B3E")
STEEL = PatternFill("solid", fgColor="1C3A6E")
WHITE = PatternFill("solid", fgColor="FFFFFF")
ALT   = PatternFill("solid", fgColor="E2EAF6")
RED   = PatternFill("solid", fgColor="FDECEA")

HDR_F   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_F  = Font(name="Arial", size=10)
TITLE_F = Font(name="Arial", bold=True, color="0D1B3E", size=13)
SUB_F   = Font(name="Arial", color="64748B", size=9)
BOLD_F  = Font(name="Arial", bold=True, color="0D1B3E", size=10)

CTR  = Alignment(horizontal="center", vertical="center")
LFT  = Alignment(horizontal="left",   vertical="center")
WRP  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="D1DAE8")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HDR_F; c.fill = NAVY; c.alignment = CTR; c.border = BDR

def bc(ws, row, col, val, align=CTR, font=None, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or BODY_F; c.alignment = align; c.border = BDR
    if fill: c.fill = fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, t1, t2, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]; c.value = t1; c.font = TITLE_F; c.alignment = LFT
    ws.merge_cells(f"A2:{get_column_letter(span)}2")
    c = ws["A2"]; c.value = t2; c.font = SUB_F; c.alignment = LFT
    ws.row_dimensions[3].height = 6


def build_summary_sheet(ws, df):
    """One row per (date × resource) — full dataset."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    title_block(ws,
        f"ERCOT ONRUC Commitments — {START_DATE} to present",
        "Source: ERCOT Public API NP1-301 (60-Day COP Snapshot)  ·  ONRUC = unit committed for RUC  ·  One row per resource per operating day",
        8)

    hdrs = ["Operating Date", "Resource", "QSE", "HE Range", "HSL (MW)", "Status", "Month", "Year"]
    for i, h in enumerate(hdrs, 1):
        hdr(ws, 4, i, h)

    # Deduplicate to one row per (date, resource) with HE range
    def he_range(nums):
        nums = sorted(set([int(n) for n in nums if n is not None]))
        if not nums: return ""
        if len(nums) == 1: return f"HE{nums[0]}"
        if nums == list(range(nums[0], nums[-1]+1)):
            return f"HE{nums[0]}–HE{nums[-1]}"
        return ", ".join(f"HE{n}" for n in nums)

    dedup = (df.groupby(["operatingdate", "resource"])
               .agg(he_range_=("he", he_range),
                    hsl=("hsl", "first"),
                    qse=("qse", "first"),
                    status=("status", "first"))
               .reset_index())
    dedup = dedup.sort_values(["operatingdate", "resource"]).reset_index(drop=True)

    for idx, row in enumerate(dedup.itertuples(index=False), 5):
        f = ALT if idx % 2 == 0 else WHITE
        ws.row_dimensions[idx].height = 18
        try:
            date_str = row.operatingdate.strftime("%-m/%-d/%Y")
            month_str = row.operatingdate.strftime("%b %Y")
            year_str  = str(row.operatingdate.year)
        except:
            date_str = month_str = year_str = str(row.operatingdate)[:10]

        bc(ws, idx, 1, date_str,          CTR, BOLD_F, f)
        bc(ws, idx, 2, row.resource,      LFT, fill=f)
        bc(ws, idx, 3, row.qse,           CTR, fill=f)
        bc(ws, idx, 4, row.he_range_,     CTR, fill=f)
        hsl_val = int(row.hsl) if pd.notna(row.hsl) else "N/A"
        bc(ws, idx, 5, hsl_val,           CTR, fill=f)
        bc(ws, idx, 6, row.status,        CTR, fill=f)
        bc(ws, idx, 7, month_str,         CTR, fill=f)
        bc(ws, idx, 8, year_str,          CTR, fill=f)

    set_widths(ws, [14, 22, 16, 16, 12, 10, 12, 8])

    # Color scale on HSL column
    last = 4 + len(dedup)
    ws.conditional_formatting.add(
        f"E5:E{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color="C0392B")
    )


def build_month_sheet(ws, df_month, month_label):
    """One row per (date × resource) for a single month."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    title_block(ws,
        f"ERCOT ONRUC Commitments — {month_label}",
        "Source: ERCOT Public API NP1-301  ·  ONRUC status = unit committed for RUC  ·  One row per resource per day",
        6)

    hdrs = ["Operating Date", "Resource", "QSE", "HE Range", "HSL (MW)", "Status"]
    for i, h in enumerate(hdrs, 1):
        hdr(ws, 4, i, h)

    def he_range(nums):
        nums = sorted(set([int(n) for n in nums if n is not None]))
        if not nums: return ""
        if len(nums) == 1: return f"HE{nums[0]}"
        if nums == list(range(nums[0], nums[-1]+1)):
            return f"HE{nums[0]}–HE{nums[-1]}"
        return ", ".join(f"HE{n}" for n in nums)

    dedup = (df_month.groupby(["operatingdate", "resource"])
                     .agg(he_range_=("he", he_range),
                          hsl=("hsl", "first"),
                          qse=("qse", "first"),
                          status=("status", "first"))
                     .reset_index()
                     .sort_values(["operatingdate", "resource"])
                     .reset_index(drop=True))

    for idx, row in enumerate(dedup.itertuples(index=False), 5):
        f = ALT if idx % 2 == 0 else WHITE
        ws.row_dimensions[idx].height = 18
        try:
            date_str = row.operatingdate.strftime("%-m/%-d/%Y")
        except:
            date_str = str(row.operatingdate)[:10]

        bc(ws, idx, 1, date_str,       CTR, BOLD_F, f)
        bc(ws, idx, 2, row.resource,   LFT, fill=f)
        bc(ws, idx, 3, row.qse,        CTR, fill=f)
        bc(ws, idx, 4, row.he_range_,  CTR, fill=f)
        hsl_val = int(row.hsl) if pd.notna(row.hsl) else "N/A"
        bc(ws, idx, 5, hsl_val,        CTR, fill=f)
        bc(ws, idx, 6, row.status,     CTR, fill=f)

    set_widths(ws, [14, 22, 16, 16, 12, 10])


def save_excel(result):
    print(f"\nBuilding Excel workbook ...")
    wb = Workbook()
    wb.remove(wb.active)

    # Summary tab
    ws_sum = wb.create_sheet("Summary")
    build_summary_sheet(ws_sum, result)
    print("  Summary tab built")

    # One tab per month, ordered chronologically
    result["month_period"] = result["operatingdate"].dt.to_period("M")
    for period in sorted(result["month_period"].unique()):
        label = period.strftime("%b %Y")
        df_m = result[result["month_period"] == period]
        ws = wb.create_sheet(label)
        build_month_sheet(ws, df_m, label)
        print(f"  {label}: {df_m['resource'].nunique()} resources, {df_m['operatingdate'].nunique()} event days")

    wb.save(OUTPUT_XLSX)
    print(f"\n{'='*60}")
    print(f"  Saved Excel: {OUTPUT_XLSX}")
    print(f"  Tabs: Summary + {len(wb.sheetnames)-1} monthly tabs")
    print(f"{'='*60}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  ERCOT COP ONRUC Pull")
    print(f"  Date range: {START_DATE} → {END_DATE}")
    print(f"  Note: data is on a 60-day posting lag")
    print("=" * 60)

    if "YOUR_EMAIL" in USERNAME or "YOUR_SUBSCRIPTION_KEY" in SUBSCRIPTION_KEY:
        print("\n⚠  CREDENTIALS NOT SET")
        print("   Export as environment variables:")
        print("     export ERCOT_USERNAME='your@email.com'")
        print("     export ERCOT_PASSWORD='yourpassword'")
        print("     export ERCOT_SUBSCRIPTION_KEY='yourkey'")
        print("\n   Register at: https://apiexplorer.ercot.com/")
        return

    print("\nStep 1: Authenticating ...")
    token = get_id_token()

    print("\nStep 2: Pulling COP data ...")
    df = fetch_all_onruc(token)

    if df.empty:
        print("No data returned.")
        return

    print("\nStep 3: Processing ...")
    result = build_result(df)

    print("\nStep 4: Saving CSV ...")
    result.to_csv(OUTPUT_CSV, index=False)
    print(f"  Saved {len(result):,} records to {OUTPUT_CSV}")

    print("\nStep 5: Building Excel ...")
    save_excel(result)

    # Quick summary
    dedup = result.drop_duplicates(subset=["operatingdate", "resource"])
    print(f"\nQuick stats:")
    print(f"  Event days:        {dedup['operatingdate'].nunique():,}")
    print(f"  Unique resources:  {dedup['resource'].nunique():,}")
    print(f"  Total records:     {len(dedup):,}")
    print(f"\nTop 10 most RUC-committed resources:")
    print(dedup["resource"].value_counts().head(10).to_string())


if __name__ == "__main__":
    main()
